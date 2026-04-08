"""AGEB census data reader for INEGI RESAGEBURB files."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

from app.models.schemas import AGEBData

logger = logging.getLogger(__name__)

# INEGI uses '*' to mark confidential (suppressed) values.
_CONFIDENTIAL = "*"

# Default area estimate per AGEB in km² when no area data is available.
# Average urban AGEB in CDMX is roughly 0.25 km².
_DEFAULT_AGEB_AREA_KM2 = 0.25

# Socioeconomic level thresholds based on average years of schooling (GRAPROES).
_NSE_THRESHOLDS: list[tuple[float, str]] = [
    (14.0, "Alto"),
    (12.0, "Medio-Alto"),
    (10.0, "Medio"),
    (7.0, "Medio-Bajo"),
    (0.0, "Bajo"),
]


def _safe_int(value: object) -> int:
    """Convert a value to int, returning 0 for confidential or non-numeric values."""
    if value is None or str(value).strip() == _CONFIDENTIAL:
        return 0
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def _safe_float(value: object) -> float:
    """Convert a value to float, returning 0.0 for confidential or non-numeric values."""
    if value is None or str(value).strip() == _CONFIDENTIAL:
        return 0.0
    try:
        return float(str(value))
    except (ValueError, TypeError):
        return 0.0


def _classify_socioeconomic(avg_schooling: float) -> str:
    """Classify socioeconomic level from average years of schooling (GRAPROES)."""
    for threshold, level in _NSE_THRESHOLDS:
        if avg_schooling >= threshold:
            return level
    return "Bajo"


def _build_ageb_id(row: pd.Series) -> str:
    """Build a 13-character AGEB ID from component columns.

    Format: ENTIDAD(2) + MUN(3) + LOC(4) + AGEB(4)
    Example: '0900200010010'
    """
    ent = str(row["ENTIDAD"]).zfill(2)
    mun = str(row["MUN"]).zfill(3)
    loc = str(row["LOC"]).zfill(4)
    ageb = str(row["AGEB"]).zfill(4)
    return f"{ent}{mun}{loc}{ageb}"


class AGEBReader:
    """Reads and queries INEGI AGEB census data from an Excel file."""

    def __init__(self) -> None:
        self._df: pd.DataFrame | None = None

    def load(self, filepath: str) -> pd.DataFrame:
        """Load the AGEB data file (CSV preferred, Excel as fallback).

        If a CSV file exists at the same path with .csv extension, it will be
        used instead of the Excel file for much faster loading.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file is corrupt or cannot be parsed.
        """
        path = Path(filepath)

        # Prefer CSV version if available (orders of magnitude faster)
        csv_path = path.with_suffix(".csv")
        if not csv_path.exists():
            csv_path = path.parent / "ageb_data.csv"

        if csv_path.exists():
            return self._load_csv(csv_path)

        if not path.exists():
            raise FileNotFoundError(
                f"Archivo AGEB no encontrado: {filepath}"
            )

        return self._load_excel(path)

    def _load_csv(self, csv_path: Path) -> pd.DataFrame:
        """Fast path: load from pre-processed CSV."""
        try:
            df = pd.read_csv(
                csv_path,
                dtype={"ENTIDAD": str, "MUN": str, "LOC": str, "AGEB": str, "MZA": str},
                encoding="utf-8",
            )
        except Exception as exc:
            raise ValueError(f"No se pudo leer el archivo CSV: {exc}") from exc

        required_cols = {"ENTIDAD", "MUN", "LOC", "AGEB", "POBTOT", "PEA", "GRAPROES"}
        missing = required_cols - set(df.columns)
        if missing:
            raise ValueError(f"Archivo AGEB CSV — columnas faltantes: {missing}")

        # Ensure string types
        for col in ["ENTIDAD", "MUN", "LOC", "AGEB"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        df["ageb_id"] = df.apply(_build_ageb_id, axis=1)
        self._df = df
        logger.info("Loaded %d AGEB records from CSV %s", len(df), csv_path)
        return df

    def _load_excel(self, path: Path) -> pd.DataFrame:
        """Slow path: load from Excel using openpyxl read_only mode."""
        # Only read the columns we actually need
        _NEEDED_COLS = ["ENTIDAD", "MUN", "LOC", "AGEB", "MZA",
                        "POBTOT", "PEA", "GRAPROES", "POCUPADA", "VIVPAR_HAB"]

        try:
            # First try reading only needed columns (much faster)
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(c) if c else "" for c in next(rows)]

            # Find column indices for needed columns
            col_indices: dict[str, int] = {}
            for col_name in _NEEDED_COLS:
                if col_name in header:
                    col_indices[col_name] = header.index(col_name)

            required_cols = {"ENTIDAD", "MUN", "LOC", "AGEB", "MZA", "POBTOT", "PEA", "GRAPROES"}
            missing = required_cols - set(col_indices.keys())
            if missing:
                wb.close()
                raise ValueError(
                    f"Archivo AGEB corrupto — columnas faltantes: {missing}"
                )

            # Read only needed columns into a list of dicts
            data_rows = []
            for row in rows:
                record = {}
                for col_name, idx in col_indices.items():
                    val = row[idx] if idx < len(row) else None
                    record[col_name] = str(val) if val is not None else ""
                data_rows.append(record)
            wb.close()

            df = pd.DataFrame(data_rows)

        except ValueError:
            raise  # Re-raise our own ValueError
        except Exception as exc:
            raise ValueError(
                f"No se pudo leer el archivo AGEB: {exc}"
            ) from exc

        # Ensure string types for geographic ID columns
        for col in ["ENTIDAD", "MUN", "LOC", "AGEB", "MZA"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        # Filter to AGEB-level totals (MZA=000, not municipality/locality totals)
        df = df[
            (df["MZA"] == "000")
            & (df["LOC"] != "0000")
            & (df["AGEB"] != "0000")
        ].copy()

        df["ageb_id"] = df.apply(_build_ageb_id, axis=1)
        self._df = df
        logger.info("Loaded %d AGEB records from %s", len(df), filepath)
        return df

    def get_zone_data(self, ageb_ids: list[str]) -> AGEBData:
        """Aggregate demographic data for the given AGEB IDs.

        Args:
            ageb_ids: List of 13-character AGEB identifiers.

        Returns:
            AGEBData with aggregated indicators.

        Raises:
            RuntimeError: If ``load()`` has not been called yet.
        """
        if self._df is None:
            raise RuntimeError(
                "Debe llamar a load() antes de get_zone_data()"
            )

        matched = self._df[self._df["ageb_id"].isin(ageb_ids)]
        ageb_count = len(matched)

        if ageb_count == 0:
            return AGEBData(
                total_population=0,
                population_density=0.0,
                economically_active_population=0,
                socioeconomic_level="Desconocido",
                ageb_count=0,
                raw_indicators={},
            )

        # --- Aggregate indicators ---
        def _col_sum(col: str) -> int:
            if col in matched.columns:
                return int(matched[col].apply(_safe_int).sum())
            return 0

        def _col_float_weighted_avg(col: str) -> float:
            if col not in matched.columns:
                return 0.0
            vals = matched[col].apply(_safe_float)
            pops = matched["POBTOT"].apply(_safe_int)
            total_w = pops.sum()
            if total_w > 0:
                return round((vals * pops).sum() / total_w, 2)
            return round(vals.mean(), 2) if len(vals) > 0 else 0.0

        total_pop = _col_sum("POBTOT")
        total_pea = _col_sum("PEA")
        female_pop = _col_sum("POBFEM")
        male_pop = _col_sum("POBMAS")
        pop_0_14 = _col_sum("POB0_14")
        pop_15_64 = _col_sum("POB15_64")
        pop_65_plus = _col_sum("POB65_MAS")
        occupied = _col_sum("POCUPADA")
        unemployed = _col_sum("PDESOCUP")
        inactive = _col_sum("PE_INAC")
        total_households = _col_sum("TOTHOG")
        total_dwellings = _col_sum("VIVTOT")
        vivpar_hab = _col_sum("VIVPAR_HAB")
        pop_health = _col_sum("PDER_SS")
        pop_no_health = _col_sum("PSINDER")

        # Population density
        estimated_area = ageb_count * _DEFAULT_AGEB_AREA_KM2
        pop_density = round(total_pop / estimated_area, 2) if estimated_area > 0 else 0.0

        # Avg schooling years (weighted)
        avg_schooling = _col_float_weighted_avg("GRAPROES")
        nse = _classify_socioeconomic(avg_schooling)

        # Avg occupants per dwelling
        ocupvivpar = _col_sum("OCUPVIVPAR")
        avg_occ = round(ocupvivpar / vivpar_hab, 2) if vivpar_hab > 0 else 0.0

        # Housing infrastructure percentages
        def _pct(col: str) -> float:
            if col not in matched.columns or vivpar_hab == 0:
                return 0.0
            return round(_col_sum(col) / vivpar_hab * 100, 1)

        pct_elec = _pct("VPH_C_ELEC")
        pct_water = _pct("VPH_AGUADV")
        pct_drain = _pct("VPH_DRENAJ")
        pct_internet = _pct("VPH_INTER")
        pct_car = _pct("VPH_AUTOM")
        pct_cell = _pct("VPH_CEL")
        pct_pc = _pct("VPH_PC")

        # Build raw indicators per AGEB
        raw: dict[str, dict] = {}
        for _, row in matched.iterrows():
            aid = row["ageb_id"]
            entry: dict = {}
            for col in matched.columns:
                if col not in ("ageb_id", "ENTIDAD", "MUN", "LOC", "AGEB", "MZA"):
                    entry[col.lower()] = _safe_int(row[col]) if col != "GRAPROES" and col != "PROM_OCUP" else _safe_float(row[col])
            raw[aid] = entry

        return AGEBData(
            total_population=total_pop,
            population_density=pop_density,
            economically_active_population=total_pea,
            socioeconomic_level=nse,
            ageb_count=ageb_count,
            female_population=female_pop,
            male_population=male_pop,
            population_0_14=pop_0_14,
            population_15_64=pop_15_64,
            population_65_plus=pop_65_plus,
            occupied_population=occupied,
            unemployed_population=unemployed,
            inactive_population=inactive,
            avg_schooling_years=avg_schooling,
            total_households=total_households,
            total_dwellings=total_dwellings,
            avg_occupants_per_dwelling=avg_occ,
            population_with_health_services=pop_health,
            population_without_health_services=pop_no_health,
            pct_with_electricity=pct_elec,
            pct_with_water=pct_water,
            pct_with_drainage=pct_drain,
            pct_with_internet=pct_internet,
            pct_with_car=pct_car,
            pct_with_cellphone=pct_cell,
            pct_with_computer=pct_pc,
            raw_indicators=raw,
        )
